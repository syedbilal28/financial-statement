
from sqlalchemy import engine
from openpyxl import Workbook
import openpyxl
import datetime
import re


import os
class ContraAsset:
    def __init__(self,Account:list,Contra:list):
        self.account = Account[0][0]
        self.value = Account[0][1]
        self.contra = Contra[0][0]
        self.contra_value = Contra[0][1]
class AccountingCycle:
    def __init__(self):
        self.dbpath = os.path.abspath("./data1.db")
        self.xlpath = os.path.abspath("./test.xlsx")
        self.engine = engine.create_engine("sqlite:///"+self.dbpath)
        self.assets_Expenses =["Cash","Accounts Receivable","Notes Receivable","Supplies","Equipment",'Rent Expense',"Salaries Expense","Drawings"]
        self.liabilities= ["Accounts Payable","Notes Payable","Unearned Revenue","Service Revenue","Capital"]
        self.accountnumbers= eval(open('accounts.txt','r').read())

    def AddingIntoTrial(self,debit,credit):
        print("Hrere3")
        conn = self.engine.connect()
        query = "Select Accountname from Accounts"
        print("hre4")
        accounts = conn.execute(query)
        print("Hre5")
        accounts= accounts.fetchall()

        if len(accounts) == 0:
            self.NewAccount(debit[0][0])
        account  =None
        print(accounts)
        for i in (debit):
            for j in accounts:
                if i[0] == j[0]:
                    account = True
                    break
                else:
                    account  = False
            if account == False:
                self.NewAccount(i[0])
        account = None
        for i in (credit):

            for j in accounts:
                print("Hereee")
                if i[0] == j[0]:
                    account = True
                    break
                else:
                    account = False
            if account == False:

                self.NewAccount(i[0])
        for i in range(len(debit)):

            query1= "SELECT debit from Accounts where Accountname = '{}' ".format(debit[i][0])
            debit_entry = conn.execute(query1)
            debit_value = debit_entry.fetchone()
            debit_value= int(debit_value[0])

            debit_value = debit_value+debit[i][1]

            conn.execute("Update Accounts set Debit = {} where Accountname='{}'  ".format(debit_value, debit[i][0]))
        for i in range(len(credit)):
            query1 = "SELECT credit from Accounts where Accountname = '{}' ".format(credit[i][0])
            credit_entry = conn.execute(query1)
            credit_value = credit_entry.fetchone()
            if credit_value == None:
                credit_value=0
            credit_value = int(credit_value[0])

            credit_value = credit_value + credit[i][1]

            conn.execute("Update Accounts set credit = {} where Accountname='{}'  ".format(credit_value, credit[i][0]))
        conn.close()

    def NewAccount(self,Account):
        conn = self.engine.connect()
        accountnumber = self.accountnumbers[Account]
        #query = ("Accounts.insert().values(Accountnumber = {},Accountname = {}")
        conn.execute("INSERT INTO Accounts (AccountNumber,Accountname,Debit,Credit) VALUES ('{}','{}','0','0')".format((accountnumber),Account))
        #conn.commit()
        conn.close()


    def Balancing(self,account):

        conn = self.engine.connect()
        query ="SELECT Debit,Credit FROM Accounts where Accountname = '{}' ".format(account)
        data = conn.execute(query)
        # query2 = "Select Credit from Accounts where Accountname = '%s'"%(account)
        # credit_entry = eng.execute(query2)
        data = data.fetchall()
        print(data)
        debit = int(data[0][0])

        # credit = credit_entry.fetchone()

        credit = int(data[0][1])
        final_value = debit-credit
        print(final_value)
        if final_value >0:
            #final_value = str(final_value)
            conn.execute("Update Accounts set Debit = {} where Accountname ='{}' ".format(final_value,account))
            conn.execute("Update Accounts set Credit = {} where Accountname ='{}' ".format(0,account))
        elif final_value < 0:
            #final_value=str(final_value)
            conn.execute("Update Accounts set Credit = {} where Accountname ='{}' ".format(abs(final_value),account))
            conn.execute("Update Accounts set Debit = {} where Accountname ='{}' ".format(0,account))
        else:
            #final_value=str(final_value)
            conn.execute("Update Accounts set Debit = {} where Accountname ='{}' ".format(0,account))
            conn.execute("Update Accounts set Credit = {} where Accountname ='{}' ".format(0,account))
        conn.close()



    def initializeSpreadSheet(self):

        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(1, 1, "Date")
        sheet.cell(1, 2, "Explanation")
        sheet.cell(1, 3, "Debit")
        sheet.cell(1, 4, "Credit")
        workbook.save(self.xlpath)
    def JournalEntry(self,debit:list,credit:list,Date):
        workbook = openpyxl.load_workbook(filename=self.xlpath)
        sheet = workbook.get_sheet_by_name("Sheet")
        #date = Date
        #print(datetime.datetime.strptime("2099-12-31","%Y-%m-%d")<=Date)
        row = 2
        x = sheet.cell(row, 1).value
        Date= Date.strftime("%Y-%m-%d")
        try:
            a = datetime.datetime.strptime(x, "%Y-%m-%d")
        except:
            a= datetime.datetime.strptime("1990-12-31","%Y-%m-%d")

        b = datetime.datetime.strptime(Date, "%Y-%m-%d")
        flag = False
        Max_row = sheet.max_row

        for i in range(Max_row):
            if a>b:
                break

            print(a)
            # if b < a :
            #   break
            row += 1
            a = sheet.cell(row, 1).value

            if a!= None:
                a = datetime.datetime.strptime(a, "%Y-%m-%d")
                prev =a
            else:
                try:
                    a = prev
                except:
                    a = datetime.datetime.strptime("1990-12-31", "%Y-%m-%d")

            flag= True

        print("hre3")
        if flag== True:
            row-=1
        else:
            row= sheet.max_row +1



        if sheet.max_row ==1:
            row=2

        for i in range(len(debit)):
            sheet.insert_rows(row)
            if i  ==0:
                sheet.cell(row,1,Date)
            column = 2
            for j in range(len(debit[0])):

                sheet.cell(row,column,debit[i][j])
                column+=1

            row+=1
        for i in range(len(credit)):
            sheet.insert_rows(row)
            column = 2
            for j in range(len(credit[0])):
                if j == 1:
                    column +=1
                sheet.cell(row, column, credit[i][j])
                column += 1

            row += 1

        workbook.save(filename=self.xlpath)



    def Ledger(self,debit,credit):
        conn = self.engine.connect()
        for i in range(len(debit)):
            number = self.accountnumbers[debit[i][0]]
            conn.execute("Insert into '{1}' (Debit,credit) Values ({0},'0') ".format(str(debit[i][1]),number))
        for i in range(len(credit)):
            number = self.accountnumbers[credit[i][0]]
            conn.execute("Insert into '{1}' (Debit,credit) values( '0',{0}) ".format(str(credit[i][1]),number))
        conn.close()
    def FetchExpenses(self):
        conn = self.engine.connect()
        expenses = conn.execute("Select accountname,debit from Accounts where accountnumber like '6%'")
        expenses = expenses.fetchall()
        expenses_1 = (conn.execute("Select accountname,debit from Accounts where accountnumber like '7%'").fetchall())
        expenses_2 = (conn.execute("Select accountname,debit from Accounts where accountnumber like '9%'").fetchall())
        conn.close()
        for i in ((expenses_1)):
            expenses.append(i)
        for i in (expenses_2):
            expenses.append(i)
        return expenses
    def FetchAccounts(self):
        conn = self.engine.connect()
        accounts = conn.execute("Select Accountname from Accounts").fetchall()
        conn.close()
        return accounts

    def Closing(self):

        conn = self.engine.connect()
        workbook = openpyxl.load_workbook(filename=self.xlpath)
        sheet = workbook.get_sheet_by_name("Closing")
        revenue = conn.execute("Select Accountname,credit from Accounts where Accountnumber like '4%' ")
        revenue = revenue.fetchall()
        sheet.cell(1, 1, "Date")
        sheet.cell(1, 2, "Explanation")
        sheet.cell(1, 3, "Debit")
        sheet.cell(1, 4, "Credit")

        row = 2
        sum_revenues = 0
        for i in range(len(revenue)):
            sum_revenues += revenue[i][1]

        first_entry = [revenue, [["Income Summary", sum_revenues]]]
        expenses = self.FetchExpenses()


        expenses_sum = 0
        for i in range(len(expenses)):
            expenses_sum += expenses[i][1]
        second_entry = [[["Income Summary", expenses_sum]], expenses]

        if sum_revenues-expenses_sum>0:

            third_entry= [[["Income Summary",sum_revenues-expenses_sum]],[["Capital",sum_revenues-expenses_sum]]]
        else:
            third_entry = [[["Capital",abs(sum_revenues-expenses_sum)]],[["Income Summary",abs(sum_revenues-expenses_sum)]]]

        drawings = conn.execute("Select accountname,debit from Accounts where accountnumber = '306'")
        drawings = drawings.fetchall()

        fourth_entry = [[["Capital", drawings[0][1]]], drawings]
        Entries=[first_entry,second_entry,third_entry,fourth_entry]
        accounts= self.FetchAccounts()

        for i in Entries:
            self.AddingIntoTrial(i[0],i[1])


        for i in range(len(accounts)):
            self.Balancing(accounts[i][0])

        conn.execute("Delete from Accounts where accountnumber> '305' ")
        for s in range(4):
            for i in range(len(Entries[s])):


                for j in range(len(Entries[s][i])):
                    print(Entries[s][i])
                    column =2
                    for k in range(len(Entries[s][i][j])):
                        if column ==5:
                            break
                        if i ==1 and k ==1:
                            column+=1
                        sheet.cell(row,column,Entries[s][i][j][k])
                        column+=1
                    row+=1
        workbook.save(filename=self.xlpath)
        conn.close()

    def FetchAssets(self):
        conn = self.engine.connect()
        assets = conn.execute("Select accountname,Debit,Credit from Accounts where accountnumber like '1%'")
        assets = assets.fetchall()
        conn.close()
        return assets
    def InitializeStatements(self,sheet):
        row = 3
        sheet.cell(row, 2, "Usman Institute Of Technology")
        row += 1
        sheet.cell(row, 2, "For the month ended " + datetime.datetime.now().strftime("%B"))
        row += 1
        return row

    def IncomeStatement(self):
        conn=self.engine.connect()
        revenue = conn.execute("Select accountname,credit from Accounts where accountnumber = '400'").fetchall()
        workbook = openpyxl.load_workbook(filename= self.xlpath)
        sheet= workbook.get_sheet_by_name("Income Statement")
        row= self.InitializeStatements(sheet)
        sheet.cell(row,1,"Revenues")
        row+=1
        total_assets=0
        for i in range(len(revenue)):
            column =1
            for j in range(len(revenue[i])):
                if type(revenue[i][j]) is int:
                    total_assets+=revenue[i][j]
                if revenue[i][j]!=0:
                    sheet.cell(row,column,revenue[i][j])
                column+=3
            row+=1
        sheet.cell(row,4,total_assets)
        sheet.cell(row,2,"Total Revenues")
        row+=1
        expenses =self.FetchExpenses()
        expenses_sum = 0
        for i in range(len(expenses)):
            expenses_sum += expenses[i][1]
        sheet.cell(row,1,"Expenses")
        row+=1
        for i in range(len(expenses)):
            column = 1
            for j in range(len(expenses[i])):

                if expenses[i][j] != 0:
                    sheet.cell(row, column, expenses[i][j])
                column += 3
            row += 1
        sheet.cell(row, 4, expenses_sum)
        sheet.cell(row, 2, "Total Expenses")
        row+=1
        sheet.cell(row, 1, "Net Income")
        sheet.cell(row, 4, total_assets-expenses_sum)
        row+=2
        sheet.cell(row, 3, "Owner's Equity")
        #liabilities = (conn.execute("Select accountname,credit from Accounts where accountnumber like '2%'").fetchall())
        Capital = (conn.execute("Select accountname,credit from Accounts where accountnumber = '301'").fetchall())
        drawing = (conn.execute("Select accountname,debit from Accounts where accountnumber = '306'").fetchall())
        Capital.append(drawing[0])
        row+=1


        row+=1
        total_capital = 0
        Capital.insert(1,("Net Income",total_assets-expenses_sum))
        print(Capital)
        for i in range(len(Capital)):
            column =1
            for j in range(len(Capital[i])):

                sheet.cell(row, column, Capital[i][j])

                if Capital[i][j] =="Capital" or Capital[i][j] ==total_assets-expenses_sum :

                    total_capital+= Capital[i][1]

                elif Capital[i][j] =="Drawings" :

                    total_capital-= Capital[i][1]
                if Capital[i][j] ==total_assets-expenses_sum:
                    row+=1
                    sheet.cell(row, 4, total_capital)

                column +=3
            row+=1
        self.capital = total_capital
        sheet.cell(row,4,total_capital)
        #total_capital +=total_liabilities


        workbook.save(self.xlpath)
    def BalanceSheet(self):
        conn = self.engine.connect()
        assets= self.FetchAssets()
        workbook = openpyxl.load_workbook(filename=self.xlpath)
        sheet = workbook.get_sheet_by_name("Balance Sheet")
        row = self.InitializeStatements(sheet)
        total_assets = 0
        print(assets)
        for i in range((len(assets)-1)):
            print(i)
            current = assets[i][0]

            next = assets[i+1][0]
            print(current,next)
            c=re.search("{}".format(current),next)
            contra=[]
            if c!= None:
                print("herer")
                contra.append(assets.pop(i))
                contra.append(assets.pop(i))
        for i in range(len(assets)):
            column = 1
            for j in range(len(assets[i])):
                if type(assets[i][j]) is int:
                    total_assets += assets[i][j]
                if assets[i][j] != 0:
                    sheet.cell(row, column, assets[i][j])
                column += 3

            row += 1
        total_equipment =0

        print(contra)
        for i in range(len(contra)):

            column = 1
            for j in range(len(contra[i])):
                print(i,j, column)
                if type(contra[i][j]) is int and i ==0:
                    total_equipment += contra[i][j]
                elif type(contra[i][j]) is int and i == 1:
                    total_equipment -= contra[i][j]
                if contra[i][j] != 0:
                    sheet.cell(row, column, contra[i][j])
                    column += 1
                if i == len(contra)-1 and j == len(contra[i])-1:
                    sheet.cell(row,4,total_equipment)
            row += 1

        sheet.cell(row, 4, total_assets +total_equipment)
        sheet.cell(row, 2, "Total Assets")
        row += 2
        liabilities = (conn.execute("Select accountname,credit from Accounts where accountnumber like '2%'").fetchall())
        sheet.cell(row,1,"Liabilities")
        row+=1
        total_liabilities=0
        for i in range(len(liabilities)):
            column =1
            for j in range(len(liabilities[i])):
                if type(liabilities[i][j]) is int:
                    total_liabilities+=liabilities[i][j]
                if liabilities[i][j]!=0:
                    sheet.cell(row,column,liabilities[i][j])
                column+=3
            row+=1
        sheet.cell(row,4,total_liabilities)
        sheet.cell(row,2,"Total Liabilities")
        row+=1
        sheet.cell(row,4,self.capital)
        sheet.cell(row, 1, "Capital")
        row+=1
        sheet.cell(row, 4, self.capital+total_liabilities)
        workbook.save(self.xlpath)

# a= AccountingCycle()
# a.Closing()
# a.BalanceSheet()
# #a.AddingIntoTrial([["Cash",14000]],[["Capital",14000]])
# a.JournalEntry([["Supplies",200]],[["Cash",200]],datetime.datetime.strptime("2020-01-07","%Y-%m-%d"))

# a.BalanceSheet()